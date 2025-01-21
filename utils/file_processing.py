from docx import Document
import csv
from openpyxl import load_workbook

def process_docx(file_path):
    """
    Process a DOCX file and return its text content.
    """
    try:
        document = Document(file_path)
        text = "".join(paragraph.text + "\n" for paragraph in document.paragraphs)
        return text
    except Exception as e:
        raise ValueError(f"Failed to process DOCX file: {e}")

def process_csv(file_path):
    """
    Process a CSV file and return its text content.
    """
    try:
        with open(file_path, mode="r", encoding="utf-8") as csv_file:
            reader = csv.reader(csv_file)
            text = "".join(", ".join(row) + "\n" for row in reader)
            return text
    except Exception as e:
        raise ValueError(f"Failed to process CSV file: {e}")

def process_xlsx(file_path):
    """
    Process an XLSX file and return its text content.
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                text += ", ".join(str(cell) if cell else "" for cell in row) + "\n"
        return text
    except Exception as e:
        raise ValueError(f"Failed to process XLSX file: {e}")
